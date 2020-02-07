import numpy as np
import matplotlib.pyplot as plt
from scipy.ndimage.filters import gaussian_filter
import matplotlib.cm as cm
import openpyxl


TRIAL_TYPES = ["Practice Front-Front", "Front-Front Condition 1M", "Front-Front Condition 2F",
               "Practice FrontSide-SideFront 1", "FrontSide-SideFront 1F", "FrontSide-SideFront 2M",
               "SideFront-FrontSide 3M", "SideFront-FrontSide 4F", "Practice Inverted 1", "Inverted 1M",
               "Inverted 1F"]


def generate_scatter_plot(excel_file, sheet_number):
    wb = openpyxl.load_workbook(f'Data\\{excel_file}.xlsm')
    worksheet = wb[f'Sheet{sheet_number}']
    fig, axs = plt.subplots(3, 3)
    trial_number = 1

    for ax in zip(axs.flatten()):
        row_start = 18
        column_start = 16 + (trial_number - 1) * 3
        x, y = get_coordinate_arrays(worksheet, row_start, column_start)
        ax.plot(x, y, 'k.', markersize=5)
        trial_number += 1
        if trial_number == 9:
            break

    fig.suptitle(f"Test subject {excel_file} - {TRIAL_TYPES[sheet_number - 1]}")
    plt.show()


def generate_heat_map(excel_file, sheet_number, trial_number):
    row_start = 18
    column_start = 16 + (trial_number - 1) * 3
    wb = openpyxl.load_workbook(f'Data\\{excel_file}.xlsm')
    worksheet = wb[f'Sheet{sheet_number}']
    x, y = get_coordinate_arrays(worksheet, row_start, column_start)

    img, extent = myplot(x, y, 32)
    plt.imshow(img, extent=extent, origin='lower', cmap=cm.jet)
    plt.show()


def myplot(x, y, s, bins=500):
    heatmap, xedges, yedges = np.histogram2d(x, y, bins=bins)
    heatmap = gaussian_filter(heatmap, sigma=s)

    # extent = [xedges[0], xedges[-1], yedges[0], yedges[-1]]
    extent = [-0.2, 1.2, -0.2, 1.2]
    return heatmap.T, extent


def get_coordinate_arrays(worksheet, rowStart, columnStart):
    x_coordinate_array = []
    y_coordinate_array = []

    for row_cells in worksheet.iter_rows(min_row=rowStart, min_col=columnStart, max_col=columnStart + 1):
        for index, cell in enumerate(row_cells):
            try:
                float(cell.value)
            except ValueError:
                break
            if index == 0:
                x_coord = float(cell.value)
                continue
            else:
                y_coord = float(cell.value)

            if x_coord <= 0.05 or y_coord <= 0.05:
                break
            x_coordinate_array.append(x_coord)
            y_coordinate_array.append(y_coord)

    return np.array(x_coordinate_array), np.array(y_coordinate_array)


if __name__ == "__main__":
    scatter_plot = False
    heat_map = True
    if scatter_plot:
        print("-----------SCATTER PLOT GENERATOR POC-----------")
        excel_file = input(
            "Enter the Excel file name from the Data folder you would like to generate scatter plots for:\n")
        sheet_number = int(input("Enter the sheet number you would like to generate scatter plots for (1-11):\n"))
        generate_scatter_plot(excel_file, sheet_number)
    elif heat_map:
        print("-----------HEAT MAP GENERATOR POC-----------")
        excel_file = input(
            "Enter the Excel file name from the Data folder you would like to generate a heat map for:\n")
        sheet_number = int(input("Enter the sheet number you would like to generate a heat map for (1-11):\n"))
        trial_number = int(input("Enter the trial number you would like to generate a heat map for (1-8):\n"))
        generate_heat_map(excel_file, sheet_number, trial_number)

