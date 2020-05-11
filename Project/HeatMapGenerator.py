import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import os
from scipy.ndimage.filters import gaussian_filter
from datetime import datetime

import IO


TRIAL_TYPES = ["Practice Front-Front", "Front-Front Condition 1M", "Front-Front Condition 2F",
               "Practice FrontSide-SideFront 1", "FrontSide-SideFront 1F", "FrontSide-SideFront 2M",
               "SideFront-FrontSide 3M", "SideFront-FrontSide 4F", "Practice Inverted 1", "Inverted 1M",
               "Inverted 1F"]


def generate_scatter_plot(wb, excel_file_name, sheet_number):
    """Generates a scatter plot containing all trials for a specific scenario."""

    worksheet = wb[f'Sheet{sheet_number}']
    fig, axs = plt.subplots(3, 3)
    trial_number = 1

    for ax in axs.flatten():
        row_start = 18
        column_start = 16 + (trial_number - 1) * 3
        success, x, y = get_coordinate_arrays(worksheet, row_start, column_start)

        if success:
            ax.plot(x, y, 'k.', markersize=20)

        trial_number += 1
        if trial_number == 9:
            break

    fig.suptitle(f"Test subject {excel_file_name} - {TRIAL_TYPES[sheet_number - 1]}")
    plt.show()


def generate_heatmap_image(wb, excel_file_name, sheet_number, trial_number):
    """Generates heatmap image of the trial being searched."""

    xmin, xmax = 0, 1
    ymin, ymax = 0, 1

    # Specifies excel row and column start of data depending on trial selected.
    row_start = 18
    column_start = 16 + (trial_number - 1) * 3
    image_filepath = f"Images\\{excel_file_name}\\{TRIAL_TYPES[sheet_number-1]}_Trial{trial_number}.png"

    worksheet = wb[f'Sheet{sheet_number}']
    success, x, y = get_coordinate_arrays(worksheet, row_start, column_start)

    # Create arrays specifying the bin edges.
    nbins = 300
    xbins = np.linspace(xmin, xmax, nbins)
    ybins = np.linspace(ymin, ymax, nbins)

    if success:
        img = generate_heatmap_data(x, y, 15, (xbins, ybins))
        plt.imshow(img, origin='lower', cmap="Greys", aspect='auto', interpolation='nearest', extent=[xmin, xmax, ymin, ymax])
        plt.axis('off')
        plt.savefig(image_filepath)
    else:
        IO.print_and_log(f"Unable to generate heatmap for {excel_file_name} - Sheet: {sheet_number} Trial: {trial_number}")


def generate_heatmap_data(x, y, s, bins):
    heatmap, xedges, yedges = np.histogram2d(x, y, bins=bins)
    heatmap = gaussian_filter(heatmap, sigma=s)
    return heatmap.T


def get_coordinate_arrays(worksheet, row_start, column_start):
    """Returns eye-tracking data as an array of x-y coordinates of the trial being searched."""

    x_coordinate_array = []
    y_coordinate_array = []

    for row_cells in worksheet.iter_rows(min_row=row_start, min_col=column_start, max_col=column_start + 1):
        for index, cell in enumerate(row_cells):
            try:
                float(cell.value)
            except ValueError:
                break
            if index == 0:
                x_coord = round(float(cell.value), 2)
                continue
            else:
                y_coord = round(float(cell.value), 2)

            if x_coord <= 0.05 or y_coord <= 0.05:
                break

            x_coordinate_array.append(x_coord)
            y_coordinate_array.append(y_coord)

    if x_coordinate_array and y_coordinate_array:
        return True, np.array(x_coordinate_array), np.array(y_coordinate_array)
    else:
        return False, x_coordinate_array, y_coordinate_array


def heatmap_cli():
    """Command-line interface for generating either a scatter-plot of all trials for a specific or generate
    heatmap images for all trials done by one participant."""

    # -----------SCATTER PLOT GENERATOR-----------
    scatter_plot = input("Would you like to generate a table of scatter plots for all trials of a specific type? Y/N\n")

    if scatter_plot.lower() == 'y':
        print("-----------SCATTER PLOT GENERATOR-----------")

        excel_file = input("Enter the Excel file name from the Data folder you would like to generate heat maps for:\n")
        sheet_number = int(input("Enter the sheet number you would like to generate scatter plots for (1-11):\n"))
        wb = openpyxl.load_workbook(f'Data\\{excel_file}.xlsm')
        generate_scatter_plot(wb, excel_file, sheet_number)

    # -------------HEAT MAP GENERATOR--------------
    heat_map = input("Would you like to generate all heatmap images for a specific test subject in the 'Images' folder\n"
                     "NOTE: Doesn't work on mac and 88 images are generated!! Y/N\n")

    if heat_map.lower() == 'y':
        IO.print_and_log("-----------HEAT MAP GENERATOR-----------")
        print("Outputs grayscale heatmaps for each sheet / trial number of the selected excel file.")

        excel_file = input("Enter the Excel file name from the Data folder you would like to generate heat maps for:\n")

        IO.print_and_log(f"[{datetime.now()}] Starting to generate heatmap files to directory {os.getcwd()}\\Images\\{excel_file}")

        wb = openpyxl.load_workbook(f'Data\\{excel_file}.xlsm')
        IO.ensure_dir(f"Images\\{excel_file}\\")

        for sheet_number in range(1, 12):
            for trial_number in range(1, 9):
                IO.print_and_log(f"Successfully generated Sheet: {sheet_number} | Trial: {trial_number}")
                generate_heatmap_image(wb, excel_file, sheet_number, trial_number)

        IO.print_and_log(f"[{datetime.now()}] Finished generating heatmap files.")


if __name__ == "__main__":

    if not os.path.isdir(IO.LOGS_FILEPATH.rsplit('/', 1)[0] + '/'):
        IO.create_log_file()

    # TODO: Add functionality to generate heatmap images without needing a command-line interface.
    heatmap_cli()
