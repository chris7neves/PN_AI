import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy.ndimage.filters import gaussian_filter
import os
import re
import seaborn as sns
from enum import Enum


import filesystem

from SQL.Data_Extract import SQLDatabase


class DataSource(Enum):
    SQL = 1
    Excel = 2


GENERATED_FROM = DataSource.Excel
if GENERATED_FROM == DataSource.SQL:
    SQL_CONNECTION = SQLDatabase(filesystem.os.environ["PNAI_SERVER"],
                                 filesystem.os.environ["PNAI_DATABASE"],
                                 filesystem.os.environ["PNAI_USERNAME"],
                                 filesystem.os.environ["PNAI_PASSWORD"])

TRIAL_TYPES = ["Practice Front-Front", "Front-Front Condition 1M", "Front-Front Condition 2F",
               "Practice FrontSide-SideFront 1", "FrontSide-SideFront 1F", "FrontSide-SideFront 2M",
               "SideFront-FrontSide 3M", "SideFront-FrontSide 4F", "Practice Inverted 1", "Inverted 1M", "Inverted 1F"]

logging = filesystem.Logging("heatmap_generator")


def generate_scatter_plot(wb, excel_file_name, sheet_number):
    """Generates a scatter plot containing all trials for a specific scenario."""

    worksheet = wb[f'Sheet{sheet_number}']
    fig, axs = plt.subplots(3, 3)
    trial_number = 1

    for ax in axs.flatten():
        row_start = 18
        column_start = 16 + (trial_number - 1) * 3
        success, x, y = get_coordinate_arrays(worksheet, row_start, column_start)

        if success:
            ax.plot(x, y, 'k.', markersize=20)

        trial_number += 1
        if trial_number == 9:
            break

    fig.suptitle(f"Test subject {excel_file_name} - {TRIAL_TYPES[sheet_number - 1]}")
    plt.show()

def generate_heatmap_image_seaborn(wb, excel_file_name, sheet_number, trial_number):
    """Generates heatmap image of the trial being searched."""

    xmin, xmax = -0.2, 1.2
    ymin, ymax = -0.2, 1.2

    # Specifies excel row and column start of data depending on trial selected.
    row_start = 18
    column_start = 16 + (trial_number - 1) * 3
    image_filepath = f"Images/{excel_file_name}/{TRIAL_TYPES[sheet_number-1]}_Trial{trial_number}.png"

    worksheet = wb[f'Sheet{sheet_number}']
    success, x, y = get_coordinate_arrays(worksheet, row_start, column_start)

    # Seaborn Heatmap


    if success:
        try:
            ax = sns.kdeplot(x, y, cmap='Greys', shade=True, shade_lowest=False)
        except:
            logging.print_and_log(
                f"Unable to generate heatmap for {excel_file_name} - Sheet: {sheet_number} Trial: {trial_number}")
            return
        ax.set_frame_on(False)
        plt.axis('off')
        fig = ax.get_figure()
        fig.savefig(image_filepath, transparent=False, bbox_inches='tight', pad_inches=0)
        plt.close()
    else:
        logging.print_and_log(f"Unable to generate heatmap for {excel_file_name} - Sheet: {sheet_number} Trial: {trial_number}")

def generate_heatmap_image(wb, excel_file_name, sheet_number, trial_number):
    """Generates heatmap image of the trial being searched."""

    xmin, xmax = -0.2, 1.2
    ymin, ymax = -0.2, 1.2
    success, x, y = False, [], []
    image_filepath = f"Images/{excel_file_name}/{TRIAL_TYPES[sheet_number - 1]}_Trial{trial_number}.png"

    # Specifies excel row and column start of data depending on trial selected.
    if GENERATED_FROM == DataSource.Excel:
        row_start = 18
        column_start = 16 + (trial_number - 1) * 3

        worksheet = wb[f'Sheet{sheet_number}']
        success, x, y = get_coordinate_arrays(worksheet, row_start, column_start)

    elif GENERATED_FROM == DataSource.SQL:
        x, y = SQL_CONNECTION.coordinate_data(excel_file_name, TRIAL_TYPES[sheet_number-1], trial_number)
        success = x and y

    # Create arrays specifying the bin edges.
    nbins = 300
    xbins = np.linspace(xmin, xmax, nbins)
    ybins = np.linspace(ymin, ymax, nbins)

    if success:
        logging.print_and_log(f"Successfully generated Sheet: {sheet_number} | Trial: {trial_number}")

        img = generate_heatmap_data(x, y, 15, (xbins, ybins))
        plt.imshow(img, origin='lower', cmap="Greys", aspect='auto', interpolation='nearest', extent=[xmin, xmax, ymin, ymax])
        plt.axis('off')
        plt.savefig(image_filepath)
        plt.close()
    else:
        logging.print_and_log(f"Unable to generate heatmap for {excel_file_name} - Sheet: {sheet_number} Trial: {trial_number}")


def generate_heatmap_data(x, y, s, bins):
    heatmap, xedges, yedges = np.histogram2d(x, y, bins=bins)
    heatmap = gaussian_filter(heatmap, sigma=s)
    return heatmap.T


def get_coordinate_arrays(worksheet, row_start, column_start):
    """Returns eye-tracking data as an array of x-y coordinates of the trial being searched."""

    x_coordinate_array = []
    y_coordinate_array = []

    for row_cells in worksheet.iter_rows(min_row=row_start, min_col=column_start, max_col=column_start + 2):
        if not str(row_cells[0].value).replace('.', '', 1).isdigit():
            continue

        x_coord = round(float(row_cells[0].value), 2)
        y_coord = round(float(row_cells[1].value), 2)
        time = round(float(row_cells[2].value), 2)

        if not 0.05 <= x_coord <= 0.95 or not 0.05 <= y_coord <= 0.95 or time < 0:
            continue

        x_coordinate_array.append(x_coord)
        y_coordinate_array.append(y_coord)

    return x_coordinate_array and y_coordinate_array, np.array(x_coordinate_array), np.array(y_coordinate_array)

def get_hm_img(hm_type, wb, file_name, sheet_number, trial_number):
    """
    Handles the selection of which heatmap is to be generated. Returns either a standard heatmap or seaborn
    heatmap.

    :param hm_type: heatmap type user wishes to generate
    :param wb: openpyxl workbook object
    :param file_name: name of file, or 'a' for all
    :param sheet_number: sheet number to be generated
    :param trial_number: trial number to be generated
    :return: returns a heatmap of the chosen type
    """
    if hm_type.lower() == 'm':
        return generate_heatmap_image(wb, file_name, sheet_number, trial_number)
    elif hm_type.lower() == 's':
        return generate_heatmap_image_seaborn(wb, file_name, sheet_number, trial_number)

def generate_heatmaps(file_name, hm_type):

    wb = None

    if file_name.lower() != 'a' and '.' not in file_name:
        file_name = file_name.upper()

    if file_name.lower() == 'a':
        cwd = os.getcwd()
        data_folder = os.path.join(cwd, "Data")
        for r, d, f in os.walk(data_folder):
            for file in f:
                if re.match('(A|T)[0-9]{3}.xls', file):

                    file_name = file.rsplit('.')[0]

                    logging.print_and_log(f"Starting to generate heatmap files to directory /images/{file_name}")
                    print(f"Generating heatmaps for {file_name} ...")
                    filesystem.ensure_filepath_exists(f"images/{file_name}/")

                    if GENERATED_FROM == DataSource.Excel:
                        wb = openpyxl.load_workbook(f'data/{file}')

                    for sheet_number in range(1, 12):
                        for trial_number in range(1, 9):
                            get_hm_img(hm_type, wb, file_name, sheet_number, trial_number)

                    if GENERATED_FROM == DataSource.Excel:
                        wb.close()

    elif re.match('(A|T)[0-9]{3}', file_name):
        logging.print_and_log(f"Starting to generate heatmap files to directory /images/{file_name}")
        print(f"Generating heatmaps for {file_name} ...")
        filesystem.ensure_filepath_exists(f"images/{file_name}/")

        if GENERATED_FROM == DataSource.Excel:
            wb = openpyxl.load_workbook(f'data/{file_name}.xlsm')

        for sheet_number in range(1, 12):
            for trial_number in range(1, 9):
                get_hm_img(hm_type, wb, file_name, sheet_number, trial_number)

        if GENERATED_FROM == DataSource.Excel:
            wb.close()


def heatmap_cli():
    """Command-line interface for generating either a scatter-plot of all trials for a specific or generate
    heatmap images for all trials done by one participant."""

    # -----------SCATTER PLOT GENERATOR-----------
    scatter_plot = input("Would you like to generate a table of scatter plots for all trials of a specific type? Y/N\n")

    if scatter_plot.lower() == 'y':
        print("-----------SCATTER PLOT GENERATOR-----------")

        excel_file = input("Enter the Excel file name from the Data folder you would like to generate heat maps for:\n")
        sheet_number = int(input("Enter the sheet number you would like to generate scatter plots for (1-11):\n"))
        wb = openpyxl.load_workbook(f'data/{excel_file}.xlsm')
        generate_scatter_plot(wb, excel_file, sheet_number)

    # -------------HEAT MAP GENERATOR--------------
    heat_map = input("Would you like to generate all heatmap images for a specific test subject in the 'Images' folder\n"
                     "NOTE: Doesn't work on mac and 88 images are generated!! Y/N\n")

    heat_map_type = input("Enter m for standard heatmap or s for seaborn heatmaps:")



    if heat_map.lower() == 'y':
        logging.print_and_log("-----------HEAT MAP GENERATOR-----------")
        print("Outputs grayscale heatmaps for each sheet / trial number of the selected excel file.")

        excel_file = input("Enter the Excel file name from the Data folder you would like to generate heatmaps for or enter 'a' to generate heatmaps for all trials in data folder:\n")

        generate_heatmaps(excel_file, heat_map_type)

    logging.print_and_log(f"Finished generating heatmap files.")


if __name__ == "__main__":

    # TODO: Add functionality to generate heatmap images without needing a command-line interface.
    heatmap_cli()
